import os
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def find_image_in_directory(directory, extensions=(".png", ".jpg", ".jpeg")):
    """
    Searches for an image in the specified directory.
    
    Args:
        directory (str): The directory to search for images.
        extensions (tuple): File extensions to look for.
        
    Returns:
        str: Path to the first image file found, or None if no image is found.
    """
    for file in os.listdir(directory):
        if file.lower().endswith(extensions):
            return os.path.join(directory, file)
    return None

def image_to_excel(output_excel, cell_size=10):
    """
    Converts the first image in the current directory to pixel art in Excel.
    
    Args:
        output_excel (str): Path to the output Excel file.
        cell_size (int): Resize image to this dimension (cell_size x cell_size).
    """
    # Find the first image in the current directory
    image_path = find_image_in_directory(os.getcwd())
    if not image_path:
        print("No image found in the current directory.")
        return

    # Open the image and resize it
    img = Image.open(image_path)
    img = img.resize((cell_size, cell_size), Image.Resampling.NEAREST)
    
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    
    # Loop through the resized image's pixels
    for row in range(img.height):
        for col in range(img.width):
            # Get pixel color (R, G, B)
            r, g, b = img.getpixel((col, row))[:3]
            # Convert RGB to HEX
            hex_color = f"{r:02X}{g:02X}{b:02X}"
            # Fill the cell with the color
            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            cell = ws.cell(row=row+1, column=col+1)
            cell.fill = fill
            # Adjust cell width/height to square shape
            ws.column_dimensions[cell.column_letter].width = 2.5
        ws.row_dimensions[row+1].height = 15
    
    # Save the Excel file
    wb.save(output_excel)
    print(f"Excel file saved at: {output_excel}")

# Example usage
image_to_excel("pixel_art.xlsx", cell_size=200)
